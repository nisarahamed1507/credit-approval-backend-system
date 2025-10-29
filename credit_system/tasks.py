from celery import shared_task
from django.db import transaction
from .models import Customer, Loan
import openpyxl
from datetime import datetime
from decimal import Decimal
import os


@shared_task
def ingest_customer_data(file_path='data/customer_data.xlsx'):
    """
    Background task to ingest customer data from Excel file
    """
    try:
        full_path = os.path.join('/app', file_path)
        
        if not os.path.exists(full_path):
            return f"File not found: {full_path}"
        
        workbook = openpyxl.load_workbook(full_path)
        sheet = workbook.active
        
        customers_created = 0
        customers_updated = 0
        
        # Skip header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
                
            customer_id = int(row[0])
            first_name = str(row[1])
            last_name = str(row[2])
            phone_number = int(row[3])
            monthly_salary = Decimal(str(row[4]))
            approved_limit = Decimal(str(row[5]))
            current_debt = Decimal(str(row[6])) if row[6] else Decimal('0.00')
            
            with transaction.atomic():
                customer, created = Customer.objects.update_or_create(
                    customer_id=customer_id,
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'phone_number': phone_number,
                        'monthly_salary': monthly_salary,
                        'approved_limit': approved_limit,
                        'current_debt': current_debt,
                        'age': 25  # Default age, as Excel doesn't have this field
                    }
                )
                
                if created:
                    customers_created += 1
                else:
                    customers_updated += 1
        
        return f"Customer data ingestion complete. Created: {customers_created}, Updated: {customers_updated}"
        
    except Exception as e:
        return f"Error ingesting customer data: {str(e)}"


@shared_task
def ingest_loan_data(file_path='data/loan_data.xlsx'):
    """
    Background task to ingest loan data from Excel file
    """
    try:
        full_path = os.path.join('/app', file_path)
        
        if not os.path.exists(full_path):
            return f"File not found: {full_path}"
        
        workbook = openpyxl.load_workbook(full_path)
        sheet = workbook.active
        
        loans_created = 0
        loans_updated = 0
        errors = []
        
        # Skip header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            try:
                customer_id = int(row[0])
                loan_id = int(row[1])
                loan_amount = Decimal(str(row[2]))
                tenure = int(row[3])
                interest_rate = Decimal(str(row[4]))
                monthly_repayment = Decimal(str(row[5]))
                emis_paid_on_time = int(row[6])
                
                # Parse dates
                start_date = row[7]
                end_date = row[8]
                
                # Convert to date objects if they're datetime
                if isinstance(start_date, datetime):
                    start_date = start_date.date()
                if isinstance(end_date, datetime):
                    end_date = end_date.date()
                
                # Get customer
                try:
                    customer = Customer.objects.get(customer_id=customer_id)
                except Customer.DoesNotExist:
                    errors.append(f"Customer {customer_id} not found for loan {loan_id}")
                    continue
                
                with transaction.atomic():
                    loan, created = Loan.objects.update_or_create(
                        loan_id=loan_id,
                        defaults={
                            'customer': customer,
                            'loan_amount': loan_amount,
                            'tenure': tenure,
                            'interest_rate': interest_rate,
                            'monthly_repayment': monthly_repayment,
                            'emis_paid_on_time': emis_paid_on_time,
                            'start_date': start_date,
                            'end_date': end_date,
                        }
                    )
                    
                    if created:
                        loans_created += 1
                    else:
                        loans_updated += 1
                        
            except Exception as e:
                errors.append(f"Error processing row: {str(e)}")
        
        result = f"Loan data ingestion complete. Created: {loans_created}, Updated: {loans_updated}"
        if errors:
            result += f"\nErrors: {len(errors)}"
        
        return result
        
    except Exception as e:
        return f"Error ingesting loan data: {str(e)}"


@shared_task
def ingest_all_data():
    """
    Master task to ingest both customer and loan data
    """
    customer_result = ingest_customer_data()
    loan_result = ingest_loan_data()
    
    return f"{customer_result}\n{loan_result}"
